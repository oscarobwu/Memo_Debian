#!/usr/bin/python3
# -*- coding: utf-8 -*-
import pymysql
import openpyxl
import pandas as pd 
import numpy as np 
import time

wb = openpyxl.load_workbook(r'D:\share\20201231\python_test\123\Singapore_KU正式區主機清單.xlsx')
ws = wb['SingaporeSite']
##
# map is a convenient way to construct a list. you can get a 2x2 tuple by slicing 
# openpyxl.worksheet.worksheet.Worksheet instance and last row of worksheet 
# from openpyxl.worksheet.worksheet.Worksheet.max_row
data = map(lambda x: {'Hostname': x[1].value, 
                      'Function': x[3].value, 
                      'VLAN': x[4].value, 
                      'Private_IP': x[5].value}, 
           ws[656: ws.max_row])
           # ws[16: ws.max_row]) 數字為開始行數 可以用於跳開前面幾行


# filter is another builtin function. Filter blank cells out if needed
# data = filter(lambda x: None not in x.values(), data)
df = pd.DataFrame(data) 
#df = df.columns.map(lambda x: x.replace('\r','').replace('\n', ''))
#df.fillna(0)
#df.fillna(df.mean(), inplace=True)
df = df.fillna("nnnnnn")

#db = MySQLdb.connect('host', 'user', 'password', 'database')
connection = pymysql.connect(host='192.168.88.13', # 数据库地址
                             user='testuser',       # 数据库用户名
                             password='test123',# 数据库密码
                             db='TESTDB')         # 数据库名称
                             #charset='utf8mb4')
                             #cursorclass=pymysql.cursors.DictCursor)
#cursor = db.cursor()
cursor = connection.cursor()
countt=0
#for row in data:
for index, row in df.iterrows():
    # row = row.columns.map(lambda x: x.replace('\r','').replace('\n', ''))
    # execute raw MySQL syntax by using execute function
    #time.sleep(1)
    #row = row.replace('\r',' -').replace('\n', '')
    #cursor.execute('INSERT INTO HOSTABLE2 (Hostname, Function, VLAN, Private_IP) VALUES ("{Hostname}", "{Function}", "{VLAN}", "{Private_IP}")'.format(**row))
    #               #.format(**row))  # construct MySQL syntax through format function
    #cursor.execute('insert into HOSTABLE2 (Hostname, Function, VLAN, Private_IP)'
    #               'values ("{Hostname}", "{Function}", "{VLAN}", "{Private_IP}")'
    #               .format(**row))  # construct MySQL syntax through format function
    h1 = row[0] 
    h2 = row[1]
    h3 = row[2]
    hs3 = str(h3).replace('\r',' _ ').replace('\n', ' - ')
    h4 = row[3].replace('\r','').replace('\n', ' - ')
    #print(row)
    #print("hostname : {Hostname} 功能 : {Function} VLAN往段 : {VLAN} IP addr : {Private_IP} ".format(**row))
    print("hostname : {} 功能 : {} VLAN往段 : {} IP addr : {} ".format(h1, h2, hs3, h4))
    # 查询数据库 inputline的条数
    select_sql = 'SELECT count(*) FROM HOSTABLE where Hostname = "{}" AND Function = "{}" AND VLAN = "{}" AND Private_IP = "{}"'.format(h1, h2, hs3, h4)
    cursor.execute(select_sql)
    count = cursor.fetchall()[0][0]
    print(count)
    if (count > 0):
        print("已经存在！")
    else:
        print("不存在，开始插入数据")
    #    print(count) 匯入資料需要注意欄位大小
    # pymysql.err.DataError: (1406, "Data too long for column 'name' at row 1")
    # 原因：  字段的长度不够存放数据
    # 解决方案： 就是更改mysql中name 字段的max_length 的长度
    # mySQL 下修改欄位長度
    # alter table CategoryItem modify column Column_Name varchar(6)
    #cursor.execute('insert into HOSTABLE2 (Hostname, Function, VLAN, Private_IP)'
    #               ' values ("{Hostname}", "{Function}", "{VLAN}", "{Private_IP}")'
    #               .format(**row))  # construct MySQL syntax through format function
        cursor.execute('insert into HOSTABLE (Hostname, Function, VLAN, Private_IP)'
                       'values ("{}", "{}", "{}", "{}")'
                       .format(h1, h2, hs3, h4))  # construct MySQL syntax through format function
    #except cursor as e:
    #    print("insert失敗："+str(e))
    #    exit()
    #print("hostname : {} 功能 : {} VLAN往段 : {} IP addr : {} ".format(row[0], row[1], row[2], row[3]))
    countt+=1
    print("Total : {}".format(countt))
    #print('insert into HOSTABLE2 (Hostname, Function, VLAN, Private_IP)'
    #               ' values ("{Hostname}", "{Function}", "{VLAN}", "{Private_IP}")'
    #               .format(**row))  # construct MySQL syntax through format function
connection.commit()

print('批量导入数据完毕')
connection.close()
