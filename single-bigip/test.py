#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       test.py
#
#        USAGE: test.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-02-09 17:47:03
#      Last modified: 2021-02-09 17:49
#     REVISION: ---
#===============================================================================
# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="./single-bigip.xlsx"
# load demo.xlsx
wb=load_workbook(filepath, read_only=True, data_only=True)
# select demo.xlsx
sheet=wb.active
# get b1 cell value
b1=sheet['B1']
# get b2 cell value
b2=sheet['B2']
# get b3 cell value
b3=sheet.cell(row=3,column=2)
# print b1, b2 and b3
print(b1.value)
print(b2.value)
print(b3.value)
