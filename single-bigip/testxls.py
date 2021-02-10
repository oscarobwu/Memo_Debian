#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       testxls.py
#
#        USAGE: testxls.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-02-10 08:07:10
#      Last modified: 2021-02-10 09:26
#     REVISION: ---
#===============================================================================
import openpyxl
import time

input_file = input("Select_flie : ")
spreadsheet = {}
wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
#for sheet in wb.get_sheet_names():
wbs = wb.sheetnames
#dict_keys = []
for sheet in wbs:
    #print(sheet)
    ansible_sheet_name = 'spreadsheet_' + sheet
    spreadsheet[ansible_sheet_name] = []
    #current_sheet = wb.get_sheet_by_name(sheet)
    current_sheet = wb[sheet]
    dict_keys = []
    #print(dict_keys)
    #print(current_sheet.max_column)
    print(ansible_sheet_name)
    #time.sleep(2)
    for c in range(1, current_sheet.max_column + 1):
        dict_keys.append(current_sheet.cell(row=1, column=c).value)
        #print(current_sheet.cell(row=1, column=c).value)
        #print(c)
        #time.sleep(2)
        #print(dict_keys)
    #print(current_sheet.max_column)
    for r in range (2, current_sheet.max_row + 1):
        temp_dict = {}
        for c in range(1, current_sheet.max_column + 1):
            temp_dict[dict_keys[c-1]] = current_sheet.cell(row=r, column=c).value
            #print(temp_dict)
        spreadsheet[ansible_sheet_name].append(temp_dict)
        print(spreadsheet[ansible_sheet_name])
        time.sleep(3)
